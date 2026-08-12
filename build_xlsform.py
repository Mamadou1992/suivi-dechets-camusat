"""Genere le XLSForm KoboToolbox pour le suivi des dechets Camusat / SONAGED."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SURVEY_COLS = ["type", "name", "label::Français (fr)", "hint::Français (fr)",
               "required", "relevant", "constraint", "constraint_message::Français (fr)",
               "calculation", "appearance", "default", "read_only", "parameters"]

SURVEY = [
    ("begin_group", "identification", "1. Identification de la collecte", "", "", "", "", "", "", "field-list", "", ""),
    ("today", "date_saisie", "Date de saisie (auto)", "", "", "", "", "", "", "", "", ""),
    ("date", "date_collecte", "Date de la collecte", "Date effective du passage", "yes", "", ". <= today()", "La date ne peut pas etre dans le futur.", "", "", "today()", ""),
    ("calculate", "client", "", "", "", "", "", "", "'camusat'", "", "", ""),
    ("select_one site", "site", "Site / Lieu de collecte", "", "yes", "", "", "", "", "minimal", "thies", ""),
    ("text", "site_autre", "Preciser le site", "", "yes", "${site} = 'autre'", "", "", "", "", "", ""),
    ("end_group", "identification", "", "", "", "", "", "", "", "", "", ""),

    ("begin_group", "interlocuteurs", "2. Interlocuteurs", "", "", "", "", "", "", "field-list", "", ""),
    ("text", "responsable_sonaged", "Responsable SONAGED", "Agent realisant la collecte", "yes", "", "", "", "", "", "", ""),
    ("text", "fonction_responsable", "Fonction du responsable", "", "", "", "", "", "", "", "Responsable QHSE", ""),
    ("text", "contact_client", "Contact chez le client", "", "yes", "", "", "", "", "", "", ""),
    ("text", "tel_contact", "Telephone du contact", "Format : 77 000 00 00", "", "", "regex(., '^[0-9 +]{9,15}$')", "Numero invalide.", "", "", "", ""),
    ("end_group", "interlocuteurs", "", "", "", "", "", "", "", "", "", ""),

    ("begin_group", "quantites", "3. Quantites collectees", "", "", "", "", "", "", "", "", ""),
    ("note", "note_qte", "Renseigner le nombre de bacs et le poids pese (kg) pour chaque flux.", "", "", "", "", "", "", "", "", ""),

    ("begin_group", "g_plastiques", "Plastiques", "", "", "", "", "", "", "field-list", "", ""),
    ("integer", "bacs_plastiques", "Nombre de bacs - Plastiques", "", "yes", "", ". >= 0 and . <= 100", "Valeur attendue entre 0 et 100.", "", "", "0", ""),
    ("decimal", "poids_plastiques", "Poids - Plastiques (kg)", "", "yes", "", ". >= 0 and . <= 5000", "Poids attendu entre 0 et 5000 kg.", "", "", "0", ""),
    ("end_group", "g_plastiques", "", "", "", "", "", "", "", "", "", ""),

    ("begin_group", "g_cartons", "Cartons", "", "", "", "", "", "", "field-list", "", ""),
    ("integer", "bacs_cartons", "Nombre de bacs - Cartons", "", "yes", "", ". >= 0 and . <= 100", "Valeur attendue entre 0 et 100.", "", "", "0", ""),
    ("decimal", "poids_cartons", "Poids - Cartons (kg)", "", "yes", "", ". >= 0 and . <= 5000", "Poids attendu entre 0 et 5000 kg.", "", "", "0", ""),
    ("end_group", "g_cartons", "", "", "", "", "", "", "", "", "", ""),

    ("begin_group", "g_autres", "Autres dechets", "", "", "", "", "", "", "field-list", "", ""),
    ("integer", "bacs_autres", "Nombre de bacs - Autres", "", "yes", "", ". >= 0 and . <= 100", "Valeur attendue entre 0 et 100.", "", "", "0", ""),
    ("decimal", "poids_autres", "Poids - Autres (kg)", "", "yes", "", ". >= 0 and . <= 5000", "Poids attendu entre 0 et 5000 kg.", "", "", "0", ""),
    ("text", "nature_autres", "Nature des autres dechets", "Ex. : DIB, bois, ferraille...", "", "${poids_autres} > 0", "", "", "", "", "", ""),
    ("end_group", "g_autres", "", "", "", "", "", "", "", "", "", ""),

    ("calculate", "total_bacs", "", "", "", "", "", "", "${bacs_plastiques} + ${bacs_cartons} + ${bacs_autres}", "", "", ""),
    ("calculate", "total_poids", "", "", "", "", "", "", "${poids_plastiques} + ${poids_cartons} + ${poids_autres}", "", "", ""),
    ("calculate", "poids_valorisable", "", "", "", "", "", "", "${poids_plastiques} + ${poids_cartons}", "", "", ""),
    ("note", "recap", "TOTAL : ${total_bacs} bac(s) - ${total_poids} kg dont ${poids_valorisable} kg valorisables.", "", "", "", "", "", "", "", "", ""),
    ("end_group", "quantites", "", "", "", "", "", "", "", "", "", ""),

    ("begin_group", "destination", "4. Destination et tracabilite", "", "", "", "", "", "", "field-list", "", ""),
    ("select_one destination", "destination_dechets", "Destination des dechets tries", "", "yes", "", "", "", "", "minimal", "site_tampon", ""),
    ("text", "site_tampon_nom", "Nom du site tampon", "", "yes", "${destination_dechets} = 'site_tampon'", "", "", "", "", "", ""),
    ("text", "num_bon", "N. du bon de collecte / pesee", "", "", "", "", "", "", "", "", ""),
    ("select_one oui_non", "levee_mensuelle", "Cette collecte fait-elle partie d'une levee mensuelle ?", "", "yes", "", "", "", "", "horizontal-compact", "non", ""),
    ("text", "num_certificat", "N. du certificat de traitement", "Delivre par le repreneur (ex. Ciments du Sahel)", "", "${levee_mensuelle} = 'oui'", "", "", "", "", "", ""),
    ("end_group", "destination", "", "", "", "", "", "", "", "", "", ""),

    ("begin_group", "preuves", "5. Preuves et validation", "", "", "", "", "", "", "field-list", "", ""),
    ("image", "photo_dechets", "Photo des dechets collectes", "", "", "", "", "", "", "", "", ""),
    ("image", "photo_bon", "Photo du bon de pesee", "", "", "", "", "", "", "", "", ""),
    ("geopoint", "gps", "Position GPS du point de collecte", "", "", "", "", "", "", "", "", ""),
    ("text", "observations", "Observations", "Anomalies, refus de tri, incidents...", "", "", "", "", "", "multiline", "", ""),
    ("image", "signature_client", "Signature du contact client", "", "", "", "", "", "", "signature", "", ""),
    ("end_group", "preuves", "", "", "", "", "", "", "", "", "", ""),
]

CHOICES_COLS = ["list_name", "name", "label::Français (fr)"]
CHOICES = [
    ("site", "thies", "Camusat Thies"),
    ("site", "dakar", "Camusat Dakar"),
    ("site", "autre", "Autre site"),
    ("destination", "site_tampon", "Site tampon (stockage temporaire)"),
    ("destination", "ciments_sahel", "Ciments du Sahel (valorisation)"),
    ("destination", "decharge", "Decharge / centre d'enfouissement"),
    ("destination", "autre", "Autre destination"),
    ("oui_non", "oui", "Oui"),
    ("oui_non", "non", "Non"),
]

SETTINGS_COLS = ["form_title", "form_id", "version", "default_language", "style"]
SETTINGS = [("Suivi des dechets tries - CAMUSAT / SONAGED",
             "suivi_dechets_camusat", "2026081202", "Français (fr)", "pages")]


def write_sheet(wb, title, cols, rows):
    ws = wb.create_sheet(title)
    ws.append(cols)
    for r in rows:
        row = list(r) + [""] * (len(cols) - len(r))
        if len(row) > 0 and row[0] == "image":
            row[len(cols) - 1] = "max-pixels=1024"
        ws.append(row)
    head_fill = PatternFill("solid", fgColor="1F6F43")
    for c in range(1, len(cols) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(c)].width = max(14, min(38, len(cols[c - 1]) + 6))
    ws.freeze_panes = "A2"
    return ws


wb = openpyxl.Workbook()
wb.remove(wb.active)
write_sheet(wb, "survey", SURVEY_COLS, SURVEY)
write_sheet(wb, "choices", CHOICES_COLS, CHOICES)
write_sheet(wb, "settings", SETTINGS_COLS, SETTINGS)
out = "/sessions/jolly-gallant-lovelace/mnt/outputs/kobo_suivi_dechets_camusat.xlsx"
wb.save(out)
print("OK", out)
